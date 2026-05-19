import asyncio
import io
import logging
import re
from datetime import datetime, timezone, timedelta
from typing import Optional

import httpx
import pandas as pd
from aiogram import Bot
from aiogram.types import BufferedInputFile
from bs4 import BeautifulSoup
from sqlalchemy import func, or_, select, update

from bot.config import settings
from bot.database.db import async_session
from bot.database.models import (
    AnalyticsSource, BroadcastMessage, Group, MessageStat, MessageStatus,
)

logger = logging.getLogger(__name__)

SYNC_TTL = 1800     # seconds between re-scrapes per record
RATE_DELAY = 0.3    # seconds between HTTP requests to t.me
MAX_REFRESH = 300   # max AnalyticsSource rows to scrape per run

_HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}


def _parse_view_count(text: str) -> int:
    """Convert Telegram view strings: '1.2K' → 1200, '3.5M' → 3500000, '42' → 42."""
    text = (text or "").strip().upper().replace("\xa0", "").replace(",", "").replace(" ", "")
    if not text:
        return 0
    try:
        if text.endswith("K"):
            return int(float(text[:-1]) * 1_000)
        if text.endswith("M"):
            return int(float(text[:-1]) * 1_000_000)
        return int(text)
    except (ValueError, AttributeError):
        return 0


class AnalyticsService:
    def __init__(self, bot: Bot) -> None:
        self._bot = bot
        self._sync_lock = asyncio.Lock()

    # ------------------------------------------------------------------ #
    #  Hub-channel view sync via web scraping                              #
    # ------------------------------------------------------------------ #

    async def sync_hub_views(self) -> None:
        """
        Scrape each hub-channel message embed page to read the current view
        count from <span class="tgme_widget_message_views">.

        Requires analytics_hub_channel_username in config.
        Skips records refreshed within SYNC_TTL seconds.
        Uses a lock to prevent concurrent runs.
        """
        hub_username = settings.analytics_hub_channel_username
        if not hub_username:
            logger.debug("analytics_hub_channel_username not set — skipping view sync")
            return
        if self._sync_lock.locked():
            logger.info("sync_hub_views already running — skipping")
            return

        async with self._sync_lock:
            now = datetime.now(timezone.utc)
            cutoff = now - timedelta(seconds=SYNC_TTL)

            async with async_session() as session:
                result = await session.execute(
                    select(AnalyticsSource)
                    .where(
                        or_(
                            AnalyticsSource.last_synced.is_(None),
                            AnalyticsSource.last_synced < cutoff,
                        )
                    )
                    .order_by(AnalyticsSource.broadcast_id.desc())
                    .limit(MAX_REFRESH)
                )
                sources = list(result.scalars().all())

            if not sources:
                return

            logger.info("Scraping hub views for %d records", len(sources))

            async with httpx.AsyncClient(
                timeout=httpx.Timeout(10.0),
                headers=_HTTP_HEADERS,
                follow_redirects=True,
            ) as client:
                for source in sources:
                    url = (
                        source.hub_message_url
                        or f"https://t.me/{hub_username}/{source.hub_message_id}"
                    )
                    views = await self._scrape_post_views(client, url)

                    update_vals: dict = {"last_synced": now}
                    if views is not None:
                        update_vals["views"] = views

                    async with async_session() as session:
                        await session.execute(
                            update(AnalyticsSource)
                            .where(AnalyticsSource.id == source.id)
                            .values(**update_vals)
                        )

                    await asyncio.sleep(RATE_DELAY)

    async def _scrape_post_views(
        self,
        client: httpx.AsyncClient,
        post_url: str,
    ) -> Optional[int]:
        """
        Fetch https://t.me/{username}/{id}?embed=1 and extract the view count
        from <span class="tgme_widget_message_views">.
        """
        embed_url = post_url if "?embed=1" in post_url else post_url + "?embed=1"
        try:
            resp = await client.get(embed_url)
            if resp.status_code == 404:
                logger.debug("Post not found: %s", embed_url)
                return None
            if resp.status_code != 200:
                logger.warning("Scrape %s → HTTP %s", embed_url, resp.status_code)
                return None
            soup = BeautifulSoup(resp.text, "html.parser")
            tag = soup.find("span", class_="tgme_widget_message_views")
            if tag:
                return _parse_view_count(tag.get_text())
            return None
        except httpx.TimeoutException:
            logger.warning("Timeout scraping %s", embed_url)
            return None
        except Exception as exc:
            logger.warning("Scrape error %s: %s", embed_url, exc)
            return None

    # ------------------------------------------------------------------ #
    #  Channel overview (for Excel caption)                                #
    # ------------------------------------------------------------------ #

    async def _scrape_channel_info(
        self,
        client: httpx.AsyncClient,
        hub_username: str,
    ) -> dict:
        """
        Scrape https://t.me/s/{username} for:
          - status  (HTTP 200 → active, 404 → not found)
          - subscribers count
          - last post ID (≈ total posts published)
        """
        url = f"https://t.me/s/{hub_username}"
        try:
            resp = await client.get(url)
            if resp.status_code == 404:
                return {"status": "Не найден", "subscribers": 0, "last_post_id": 0}
            if resp.status_code != 200:
                return {"status": "Неизвестен", "subscribers": 0, "last_post_id": 0}

            soup = BeautifulSoup(resp.text, "html.parser")

            # Subscriber count — "tgme_page_extra" usually reads "1 234 subscribers"
            subscribers = 0
            extra = soup.find(class_="tgme_page_extra")
            if extra:
                m = re.search(r"([\d][\d\s.,]*[KMkm]?)", extra.get_text().replace("\xa0", " "))
                if m:
                    subscribers = _parse_view_count(m.group(1).replace(" ", ""))

            # Last message ID from post links in the feed
            last_id = 0
            pattern = re.compile(rf"t\.me/{re.escape(hub_username)}/(\d+)", re.IGNORECASE)
            for a in soup.find_all("a", href=True):
                m = pattern.search(a["href"])
                if m:
                    last_id = max(last_id, int(m.group(1)))

            return {"status": "Активен", "subscribers": subscribers, "last_post_id": last_id}
        except Exception as exc:
            logger.warning("Channel info scrape error for %s: %s", hub_username, exc)
            return {"status": "Неизвестен", "subscribers": 0, "last_post_id": 0}

    # ------------------------------------------------------------------ #
    #  Daily summary                                                       #
    # ------------------------------------------------------------------ #

    async def send_daily_summary(self) -> None:
        async with async_session() as session:
            since = datetime.now(timezone.utc) - timedelta(hours=24)

            posts_result = await session.execute(
                select(func.count(BroadcastMessage.id)).where(
                    BroadcastMessage.status == MessageStatus.sent,
                    BroadcastMessage.created_at >= since,
                )
            )
            total_posts: int = posts_result.scalar_one() or 0

            views_result = await session.execute(
                select(func.sum(AnalyticsSource.views)).where(
                    AnalyticsSource.last_synced >= since
                )
            )
            total_views: int = views_result.scalar_one() or 0

            active_result = await session.execute(
                select(func.count(Group.id)).where(Group.is_active == True)
            )
            active_groups: int = active_result.scalar_one() or 0

        text = (
            "📊 <b>Ежедневная сводка</b>\n\n"
            f"🗓 Дата: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}\n"
            f"📨 Постов за 24ч: <b>{total_posts}</b>\n"
            f"👁 Просмотров в группах: <b>{total_views:,}</b>\n"
            f"✅ Активных групп: <b>{active_groups}</b>"
        )

        try:
            await self._bot.send_message(
                chat_id=settings.control_panel_chat_id,
                text=text,
                parse_mode="HTML",
            )
        except Exception:
            logger.exception("Failed to send daily summary")

    # ------------------------------------------------------------------ #
    #  Excel report                                                        #
    # ------------------------------------------------------------------ #

    async def send_excel_report(self) -> None:
        try:
            await self.sync_hub_views()
        except Exception:
            logger.warning("sync_hub_views failed — report uses cached values")

        # Scrape channel overview for the report caption
        hub_username = settings.analytics_hub_channel_username
        channel_info: dict = {}
        if hub_username:
            try:
                async with httpx.AsyncClient(
                    timeout=httpx.Timeout(10.0),
                    headers=_HTTP_HEADERS,
                    follow_redirects=True,
                ) as client:
                    channel_info = await self._scrape_channel_info(client, hub_username)
            except Exception:
                logger.warning("Failed to scrape channel overview for report caption")

        df_groups = await self._build_dataframe()
        df_posts = await self._build_posts_dataframe()
        df_details = await self._build_post_groups_dataframe()

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            # Sheet 1: per-group totals
            df_groups.to_excel(writer, index=False, sheet_name="По группам")
            ws1 = writer.sheets["По группам"]
            _autofit_columns(ws1)
            _style_totals_row(ws1)

            # Sheet 2: last 20 broadcasts summary
            if not df_posts.empty:
                df_posts.to_excel(writer, index=False, sheet_name="Последние 20 постов")
                ws2 = writer.sheets["Последние 20 постов"]
                _autofit_columns(ws2)

            # Sheet 3: per-post, per-group view breakdown
            if not df_details.empty:
                df_details.to_excel(writer, index=False, sheet_name="Детали по группам")
                ws3 = writer.sheets["Детали по группам"]
                _autofit_columns(ws3)

        buf.seek(0)
        filename = f"report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
        file = BufferedInputFile(buf.read(), filename=filename)

        caption = "📊 <b>Отчёт по группам</b>"
        if channel_info:
            caption += (
                f"\n\n📡 Хаб-канал: @{hub_username}\n"
                f"👥 Подписчиков: <b>{channel_info['subscribers']:,}</b>\n"
                f"📝 Постов в канале: <b>{channel_info['last_post_id']}</b>\n"
                f"📶 Статус: {channel_info['status']}"
            )

        try:
            await self._bot.send_document(
                chat_id=settings.control_panel_chat_id,
                document=file,
                caption=caption,
                parse_mode="HTML",
            )
        except Exception:
            logger.exception("Failed to send Excel report")

    # ------------------------------------------------------------------ #
    #  DataFrame builders                                                  #
    # ------------------------------------------------------------------ #

    async def _build_dataframe(self) -> pd.DataFrame:
        """Per-group aggregated stats."""
        async with async_session() as session:
            groups_result = await session.execute(select(Group))
            groups = list(groups_result.scalars().all())

            rows = []
            for group in groups:
                views_result = await session.execute(
                    select(func.sum(AnalyticsSource.views)).where(
                        AnalyticsSource.group_id == group.chat_id
                    )
                )
                total_views: int = views_result.scalar_one() or 0

                posts_result = await session.execute(
                    select(func.count(MessageStat.id)).where(
                        MessageStat.group_id == group.chat_id
                    )
                )
                post_count: int = posts_result.scalar_one() or 0

                rows.append({
                    "Название группы":   group.title,
                    "ID группы":         str(group.chat_id),
                    "Просмотры":         total_views,
                    "Доставлено постов": post_count,
                    "Статус":            "Активна" if group.is_active else "Удалена",
                })

        df = pd.DataFrame(rows)
        if df.empty:
            return df

        df = df.sort_values("Просмотры", ascending=False).reset_index(drop=True)

        total_row = {
            "Название группы":   "ИТОГО",
            "ID группы":         "",
            "Просмотры":         int(df["Просмотры"].sum()),
            "Доставлено постов": int(df["Доставлено постов"].sum()),
            "Статус":            "",
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        return df

    async def _build_posts_dataframe(self) -> pd.DataFrame:
        """Last 20 sent broadcasts with total views across all groups."""
        async with async_session() as session:
            bm_result = await session.execute(
                select(BroadcastMessage)
                .where(BroadcastMessage.status == MessageStatus.sent)
                .order_by(BroadcastMessage.created_at.desc())
                .limit(20)
            )
            broadcasts = list(bm_result.scalars().all())

            rows = []
            for bm in broadcasts:
                views_result = await session.execute(
                    select(func.sum(AnalyticsSource.views)).where(
                        AnalyticsSource.broadcast_id == bm.id
                    )
                )
                total_views: int = views_result.scalar_one() or 0

                groups_result = await session.execute(
                    select(func.count(MessageStat.id)).where(
                        MessageStat.broadcast_id == bm.id
                    )
                )
                groups_count: int = groups_result.scalar_one() or 0

                rows.append({
                    "ID поста":        bm.id,
                    "Дата":            bm.created_at.strftime("%Y-%m-%d %H:%M") if bm.created_at else "",
                    "Превью":          (bm.content_preview or "")[:120],
                    "Просмотры":       total_views,
                    "Групп получили":  groups_count,
                })

        return pd.DataFrame(rows)

    async def _build_post_groups_dataframe(self) -> pd.DataFrame:
        """Per-post, per-group view breakdown for the last 20 sent broadcasts."""
        async with async_session() as session:
            bm_result = await session.execute(
                select(BroadcastMessage)
                .where(BroadcastMessage.status == MessageStatus.sent)
                .order_by(BroadcastMessage.created_at.desc())
                .limit(20)
            )
            broadcasts = list(bm_result.scalars().all())
            if not broadcasts:
                return pd.DataFrame()

            broadcast_ids = [bm.id for bm in broadcasts]
            broadcast_map = {bm.id: bm for bm in broadcasts}

            src_result = await session.execute(
                select(AnalyticsSource)
                .where(AnalyticsSource.broadcast_id.in_(broadcast_ids))
                .order_by(AnalyticsSource.broadcast_id.desc(), AnalyticsSource.views.desc())
            )
            sources = list(src_result.scalars().all())

            group_ids = list({s.group_id for s in sources if s.group_id})
            group_map: dict = {}
            if group_ids:
                grp_result = await session.execute(
                    select(Group).where(Group.chat_id.in_(group_ids))
                )
                group_map = {g.chat_id: g.title for g in grp_result.scalars().all()}

        rows = []
        for source in sources:
            bm = broadcast_map.get(source.broadcast_id)
            group_name = group_map.get(source.group_id, str(source.group_id)) if source.group_id else "—"
            rows.append({
                "ID поста":  source.broadcast_id,
                "Дата":      bm.created_at.strftime("%Y-%m-%d %H:%M") if bm and bm.created_at else "",
                "Превью":    (bm.content_preview or "")[:80] if bm else "",
                "Группа":    group_name,
                "Просмотры": source.views or 0,
            })

        return pd.DataFrame(rows)


# ------------------------------------------------------------------ #
#  Excel helpers                                                       #
# ------------------------------------------------------------------ #

def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def _style_totals_row(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
